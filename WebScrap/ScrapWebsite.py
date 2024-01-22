
from bs4 import BeautifulSoup
import requests
import openpyxl
import time

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'All Provider Details'
print(excel.sheetnames)
sheet.append(['Company Name','Required Skills','More info'])

print("Put some skills that you are not familiar with")
unfamilar_Skills = input('>')
print(f"Filtering out {unfamilar_Skills}")

def Find_Jobs():
    html_text = requests.get('https://www.timesjobs.com/candidate/job-search.html?searchType=personalizedSearch&from=submit&txtKeywords=Python&txtLocation=').text
    soup = BeautifulSoup(html_text,"lxml")
    jobs = soup.find_all("li",class_="clearfix job-bx wht-shd-bx")

    for index, job in enumerate (jobs):
        published_date = job.find("span",class_="sim-posted").span.text

        if 'few' in published_date:
            Company_Name = job.find("h3",class_="joblist-comp-name").text.replace('  ','')
            skills = job.find("span",class_="srp-skills").text.replace(' ','')
            more_info = job.header.h2.a['href']
            if unfamilar_Skills not in skills:
                with open(f"posts/{index}.txt","w") as f:
                    f.write(f"Company Name : {Company_Name.strip()}")
                    f.write(f"Required Skills: {skills.strip()}")
                    f.write(f"More Info: {more_info}")
                print(f"File Saved : {index}")

                print(f'''Doctor Name: {Company_Name}  Title: {skills}  Expertiese: {more_info}''')

                sheet.append([Company_Name,skills,more_info])

                excel.save("Provider Details.xlsx")

if __name__=="__main__":
    while True:
        Find_Jobs()
        time_wait = 10
        print(f"Wating {time_wait} Minutes...")
        time.sleep(time_wait * 60)